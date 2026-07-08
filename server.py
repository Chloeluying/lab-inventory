"""
实验室样本出入库管理系统 - 服务器版
使用文件SQLite存储，多人共享
"""
import os
import sqlite3
import json
from datetime import datetime
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

app = Flask(__name__)
CORS(app)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'lab_data.db')


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS samples (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            样品分子号 TEXT UNIQUE NOT NULL,
            患者姓名 TEXT DEFAULT '',
            患者性别 TEXT DEFAULT '',
            患者年龄 INTEGER DEFAULT 0,
            诊断 TEXT DEFAULT '',
            样本种类 TEXT DEFAULT '',
            冰箱编号 TEXT DEFAULT '',
            冰箱层 TEXT DEFAULT '',
            收纳盒编号 TEXT DEFAULT '',
            格子序号 TEXT DEFAULT '',
            入库管数 INTEGER DEFAULT 0,
            当前余量 INTEGER DEFAULT 0,
            库位状态 TEXT DEFAULT '',
            操作日期 TEXT DEFAULT '',
            补充说明 TEXT DEFAULT '',
            登记人 TEXT DEFAULT '',
            创建时间 TEXT DEFAULT '',
            更新时间 TEXT DEFAULT ''
        )
    ''')
    conn.execute('''
        CREATE TABLE IF NOT EXISTS operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            样品分子号 TEXT NOT NULL,
            操作类型 TEXT NOT NULL,
            操作数量 INTEGER DEFAULT 0,
            操作人 TEXT DEFAULT '',
            操作日期 TEXT DEFAULT '',
            创建时间 TEXT DEFAULT '',
            FOREIGN KEY (样品分子号) REFERENCES samples(样品分子号)
        )
    ''')
    conn.commit()
    conn.close()


init_db()


def sample_from_row(row):
    if row is None:
        return None
    d = dict(row)
    # 取关联的操作记录
    conn = get_db()
    outs = conn.execute(
        'SELECT id,操作数量,操作人,操作日期 FROM operations WHERE 样品分子号=? AND 操作类型="取用" ORDER BY id',
        (d['样品分子号'],)
    ).fetchall()
    rets = conn.execute(
        'SELECT id,操作数量,操作人,操作日期 FROM operations WHERE 样品分子号=? AND 操作类型="归还" ORDER BY id',
        (d['样品分子号'],)
    ).fetchall()
    conn.close()

    d['取用记录'] = [{'id': r['id'], 'count': r['操作数量'], 'person': r['操作人'], 'date': r['操作日期']} for r in outs]
    d['归还记录'] = [{'id': r['id'], 'count': r['操作数量'], 'person': r['操作人'], 'date': r['操作日期']} for r in rets]
    # 保证旧前端兼容
    d['冰箱层（从上往下）'] = d['冰箱层']
    return d


# ===== API =====

@app.route('/')
def serve_index():
    return send_from_directory('templates', 'index.html')


@app.route('/api/samples', methods=['GET'])
def get_samples():
    conn = get_db()
    rows = conn.execute('SELECT * FROM samples ORDER BY id').fetchall()
    conn.close()
    samples = [sample_from_row(r) for r in rows]
    return jsonify({"samples": samples})


@app.route('/api/inbound', methods=['POST'])
def add_inbound():
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供数据"}), 400

    sample_id = data.get('样品分子号', '').strip()
    name = data.get('患者姓名', '').strip()
    if not sample_id or not name:
        return jsonify({"error": "缺少必填字段"}), 400

    count = int(data.get('入库管数', 1))
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    operator = data.get('登记人', '').strip()

    conn = get_db()
    existing = conn.execute('SELECT * FROM samples WHERE 样品分子号=?', (sample_id,)).fetchone()

    if existing:
        old_total = existing['入库管数']
        old_remaining = existing['当前余量']
        new_total = old_total + count
        new_remaining = old_remaining + count
        conn.execute('''UPDATE samples SET
            入库管数=?, 当前余量=?, 库位状态=?, 更新人=?, 更新时间=?
            WHERE 样品分子号=?''',
            (new_total, new_remaining, '有样本' if new_remaining > 0 else '全部取走',
             operator, now, sample_id))
    else:
        conn.execute('''INSERT INTO samples
            (样品分子号, 患者姓名, 患者性别, 患者年龄, 诊断, 样本种类,
             冰箱编号, 冰箱层, 收纳盒编号, 格子序号,
             入库管数, 当前余量, 库位状态, 操作日期, 补充说明, 登记人, 创建时间)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (sample_id, name, data.get('患者性别', ''), int(data.get('患者年龄', 0) or 0),
             data.get('诊断', ''), data.get('样本种类', ''),
             data.get('冰箱编号', ''), data.get('冰箱层', ''),
             data.get('收纳盒编号', ''), data.get('格子序号', ''),
             count, count, '有样本', data.get('操作日期', ''), data.get('补充说明', ''), operator, now))

    # 写入操作记录
    conn.execute('''INSERT INTO operations
        (样品分子号, 操作类型, 操作数量, 操作人, 操作日期, 创建时间)
        VALUES (?,?,?,?,?,?)''',
        (sample_id, '入库', count, operator, data.get('操作日期', ''), now))

    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "入库成功"})


@app.route('/api/outbound', methods=['POST'])
def add_outbound():
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供数据"}), 400

    sample_id = data.get('样品分子号', '').strip()
    count = int(data.get('操作数量', 1))

    conn = get_db()
    existing = conn.execute('SELECT * FROM samples WHERE 样品分子号=?', (sample_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "未找到该样本"}), 404

    remaining = existing['当前余量']
    total = existing['入库管数']

    if remaining < count:
        conn.close()
        return jsonify({"error": f"库存不足！当前余量 {remaining} 管"}), 400

    new_remaining = remaining - count
    if new_remaining <= 0:
        status = '全部取走'
    elif new_remaining < total:
        status = '部分取走'
    else:
        status = '有样本'

    operator = data.get('操作人', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn.execute('''UPDATE samples SET 当前余量=?, 库位状态=?, 更新时间=? WHERE 样品分子号=?''',
        (new_remaining, status, now, sample_id))

    conn.execute('''INSERT INTO operations
        (样品分子号, 操作类型, 操作数量, 操作人, 操作日期, 创建时间)
        VALUES (?,?,?,?,?,?)''',
        (sample_id, '取用', count, operator, data.get('操作日期', ''), now))

    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "取用成功"})


@app.route('/api/return', methods=['POST'])
def add_return():
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供数据"}), 400

    sample_id = data.get('样品分子号', '').strip()
    count = int(data.get('操作数量', 1))

    conn = get_db()
    existing = conn.execute('SELECT * FROM samples WHERE 样品分子号=?', (sample_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "未找到该样本"}), 404

    old_remaining = existing['当前余量']
    total = existing['入库管数']
    taken = total - old_remaining  # 实际已被取走的管数
    if count > taken:
        conn.close()
        return jsonify({"error": f"归还数不能超过已被取走的数量！该样本共被取走 {taken} 管，已归还 {old_remaining - (total - taken)} 管，最多还能还 {taken} 管"}), 400

    new_remaining = old_remaining + count

    operator = data.get('操作人', '').strip()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn.execute('''UPDATE samples SET 当前余量=?, 库位状态='有样本', 更新时间=? WHERE 样品分子号=?''',
        (new_remaining, now, sample_id))

    conn.execute('''INSERT INTO operations
        (样品分子号, 操作类型, 操作数量, 操作人, 操作日期, 创建时间)
        VALUES (?,?,?,?,?,?)''',
        (sample_id, '归还', count, operator, data.get('操作日期', ''), now))

    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "归还成功"})


@app.route('/api/delete/record', methods=['POST'])
def delete_record():
    """删除单条操作记录（取用或归还）"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供数据"}), 400

    record_id = data.get('record_id')
    if not record_id:
        return jsonify({"error": "缺少record_id"}), 400

    conn = get_db()
    rec = conn.execute('SELECT * FROM operations WHERE id=?', (record_id,)).fetchone()
    if not rec:
        conn.close()
        return jsonify({"error": "记录不存在"}), 404

    sample_id = rec['样品分子号']
    count = rec['操作数量']
    op_type = rec['操作类型']

    # 删除操作记录
    conn.execute('DELETE FROM operations WHERE id=?', (record_id,))

    # 重算样本余量
    sample = conn.execute('SELECT * FROM samples WHERE 样品分子号=?', (sample_id,)).fetchone()
    if sample:
        remaining = sample['当前余量']
        total = sample['入库管数']

        if op_type == '取用':
            # 归还被删除的取用记录
            new_remaining = min(remaining + count, total)
        else:
            # 扣回被删除的归还记录
            new_remaining = max(remaining - count, 0)

        if new_remaining <= 0:
            status = '已取完'
        elif new_remaining < total:
            status = '部分取走'
        else:
            status = '有样本'

        conn.execute('''UPDATE samples SET 当前余量=?, 库位状态=?, 更新时间=? WHERE 样品分子号=?''',
            (new_remaining, status, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), sample_id))

    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "删除成功"})


@app.route('/api/delete/sample', methods=['POST'])
def delete_sample():
    """删除整条样本（包括所有操作记录）"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "请提供数据"}), 400

    sample_id = data.get('样品分子号')
    if not sample_id:
        return jsonify({"error": "缺少样品分子号"}), 400

    conn = get_db()
    conn.execute('DELETE FROM operations WHERE 样品分子号=?', (sample_id,))
    conn.execute('DELETE FROM samples WHERE 样品分子号=?', (sample_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "success", "message": "删除成功"})


@app.route('/api/export/excel')
def export_excel():
    conn = get_db()
    rows = conn.execute('SELECT * FROM samples ORDER BY id').fetchall()
    conn.close()
    return build_excel_response(rows, f"实验室样本库存_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


@app.route('/api/export/excel/filtered')
def export_filtered_excel():
    keyword = request.args.get('keyword', '').strip()
    fridge = request.args.get('fridge', '').strip()

    conn = get_db()
    if keyword and fridge:
        rows = conn.execute(
            'SELECT * FROM samples WHERE (患者姓名 LIKE ? OR 样品分子号 LIKE ? OR 诊断 LIKE ?) AND 冰箱编号=? ORDER BY id',
            (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', fridge)
        ).fetchall()
    elif keyword:
        rows = conn.execute(
            'SELECT * FROM samples WHERE 患者姓名 LIKE ? OR 样品分子号 LIKE ? OR 诊断 LIKE ? ORDER BY id',
            (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%')
        ).fetchall()
    elif fridge:
        rows = conn.execute(
            'SELECT * FROM samples WHERE 冰箱编号=? ORDER BY id',
            (fridge,)
        ).fetchall()
    else:
        rows = conn.execute('SELECT * FROM samples ORDER BY id').fetchall()
    conn.close()

    return build_excel_response(rows, f"实验室样本库存_筛选_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


def build_excel_response(rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "样本库存"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['样品分子号', '患者姓名', '患者性别', '患者年龄', '诊断', '样本种类',
               '冰箱编号', '冰箱层', '收纳盒编号', '格子序号', '入库管数', '当前余量',
               '库位状态', '登记人', '操作日期', '补充说明']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, r in enumerate(rows, 2):
        values = [
            r['样品分子号'], r['患者姓名'], r['患者性别'], r['患者年龄'],
            r['诊断'], r['样本种类'], r['冰箱编号'], r['冰箱层'],
            r['收纳盒编号'], r['格子序号'], r['入库管数'], r['当前余量'],
            r['库位状态'], r['登记人'], r['操作日期'], r['补充说明']
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5051))
    app.run(host='0.0.0.0', port=port, debug=False)
